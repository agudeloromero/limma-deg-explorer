# ============================================================
# app.py  —  Limma DEG Interactive Explorer (Python Shiny)
#
# Author:      Patricia Agudelo-Romero
# Email:       patricia.agudeloromero@thekids.org.au
# Institution: The Kids Research Institute Australia
#
# Rules implemented in this file:
#   Rule 2  — Framework chosen for Python ecosystem fit
#   Rule 3  — Reactive expressions (@reactive.calc) separate data from display
#   Rule 4  — Input validation at the boundary (raw_data reactive)
#   Rule 5  — Every threshold is a UI control (sliders / numeric inputs)
#   Rule 6  — Analysis state captured in downloadable ZIP folder
#   Rule 9  — App deployed to URL (see 02_deploy.py)
#   Rule 10 — UI surfaces decisions, not implementation details
#
# Download output (Rule 6):
#   DEG_analysis_TIMESTAMP.zip containing:
#     ├── DEG_report_TIMESTAMP.html     (figures only, Jinja2)
#     ├── DEG_results_TIMESTAMP.csv     (full DEG table)
#     └── GO_enrichment_TIMESTAMP.xlsx  (GO_BP / GO_MF / GO_CC sheets)
#
# Key differences from R implementation (see Rule 2):
#   ┌─────────────────────┬────────────────┬──────────────────────────┐
#   │ Feature             │ R              │ Python                   │
#   ├─────────────────────┼────────────────┼──────────────────────────┤
#   │ Plotting            │ ggplot2        │ matplotlib               │
#   │ Gene labels         │ ggrepel        │ adjustText (approximate) │
#   │ GO enrichment       │ clusterProfiler│ gseapy (Enrichr API)     │
#   │ GO gene sets (mouse)│ org.Mm.eg.db   │ Enrichr mouse gene sets  │
#   │ Internet required   │ No (offline)   │ Yes (Enrichr API)        │
#   │ Report engine       │ R Markdown     │ Jinja2 + base64 figures  │
#   │ Excel output        │ openxlsx       │ openpyxl                 │
#   │ Validation syntax   │ validate()/req │ explicit None checks     │
#   │ Input access        │ input$name     │ input.name()             │
#   │ Reactive layer      │ reactive()     │ @reactive.calc           │
#   └─────────────────────┴────────────────┴──────────────────────────┘
#
# Run:
#   shiny run app.py --reload
# ============================================================

import io
import base64
import zipfile
import os
import tempfile
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import gseapy as gp
from jinja2 import Template
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from shiny import App, ui, render, reactive, req


# ── Helpers ───────────────────────────────────────────────────

def _fig_to_b64(fig):
    """Convert a matplotlib figure to a base64-encoded PNG string."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode()


def _run_go(gene_symbols, gene_set, pval_cutoff):
    """
    Run GO enrichment for one ontology via gseapy.enrichr().

    Difference from R: clusterProfiler uses org.Mm.eg.db locally
    (no internet). gseapy queries the Enrichr web API and requires
    an active internet connection.

    Column mapping (gseapy → display):
      Term             → Description
      Overlap          → Count (parsed from "N/M" string)
      P-value          → pvalue
      Adjusted P-value → p.adjust
      Genes            → geneID
    """
    if not gene_symbols:
        return None
    try:
        enr = gp.enrichr(
            gene_list = gene_symbols,
            gene_sets = gene_set,
            outdir    = None,
        )
        df = enr.results.copy()
        if df.empty:
            return None

        # Normalise column names to match R/clusterProfiler convention
        df = df.rename(columns={
            "Term":             "Description",
            "P-value":          "pvalue",
            "Adjusted P-value": "p.adjust",
            "Genes":            "geneID",
            "Combined Score":   "CombinedScore",
        })

        # Parse Count from Overlap ("5/100" → 5)
        if "Overlap" in df.columns:
            df["Count"] = df["Overlap"].str.split("/").str[0].astype(int)

        # Apply adj. P-value filter
        df = df[df["p.adjust"] <= pval_cutoff].copy()
        return df if not df.empty else None

    except Exception:
        return None


# ── GO gene set names for mouse ──────────────────────────────
# Difference from R: clusterProfiler uses ont = "BP"/"MF"/"CC"
# with org.Mm.eg.db. gseapy uses Enrichr gene set names.
GO_SETS = {
    "GO_BP": "GO_Biological_Process_2023",
    "GO_MF": "GO_Molecular_Function_2023",
    "GO_CC": "GO_Cellular_Component_2023",
}


# ── UI ───────────────────────────────────────────────────────
# Rule 5: Every analytical parameter the user might want to change
# is exposed as a slider or numeric input — never hard-coded.
# Rule 10: Labels describe decisions, not implementation details.

app_ui = ui.page_fluid(

    ui.panel_title("Limma DEG Interactive Explorer"),

    ui.tags.head(ui.tags.style("""
        body { font-family: Arial, sans-serif; }
        .num-box { margin-top: 22px; }
        .sidebar { background: #f8f9fa; }
    """)),

    ui.layout_sidebar(

        ui.sidebar(

            # ── File upload (Rule 4) ─────────────────────────
            ui.input_file("file", "Upload limma output (CSV)",
                          accept=[".csv"]),

            ui.hr(),

            # ── DEG filters (Rule 5) ─────────────────────────
            # Each filter has a slider (drag) and a numeric input (type).
            # Synchronised by @reactive.effect observers in the server.
            # Difference from R: Python uses ui.layout_columns() instead
            # of fluidRow() + column().
            ui.tags.h5("DEG Filters",
                       style="color:#2E74B5; font-weight:bold;"),

            ui.layout_columns(
                ui.input_slider("deg_pval", "Adjusted P-value \u2264",
                                min=0.001, max=1, value=0.05, step=0.001),
                ui.div({"class": "num-box"},
                       ui.input_numeric("deg_pval_num", None,
                                        min=0.001, max=1,
                                        value=0.05, step=0.001)),
                col_widths=[8, 4]
            ),

            ui.layout_columns(
                ui.input_slider("deg_lfc", "| log\u2082FC | \u2265",
                                min=0, max=4, value=1.0, step=0.1),
                ui.div({"class": "num-box"},
                       ui.input_numeric("deg_lfc_num", None,
                                        min=0, max=4,
                                        value=1.0, step=0.1)),
                col_widths=[8, 4]
            ),

            ui.hr(),

            # ── Enrichment filters (Rule 5) ──────────────────
            # Difference from R: selectInput uses a dict for choices.
            # The display tab shows one ontology; all three are
            # in the downloaded report — same behaviour as R.
            ui.tags.h5("Enrichment Filters",
                       style="color:#1E8449; font-weight:bold;"),

            ui.input_select("geneset", "Gene set (interactive preview):", {
                "GO_BP": "GO Biological Process",
                "GO_MF": "GO Molecular Function",
                "GO_CC": "GO Cellular Component",
            }),

            ui.tags.p(
                "All three ontologies are included in the downloaded report. "
                "Internet connection required.",
                style="font-size:11px; color:#888; margin-top:-6px;"
            ),

            ui.layout_columns(
                ui.input_slider("enr_pval", "Enrichment adj. P-value \u2264",
                                min=0.001, max=1, value=0.05, step=0.001),
                ui.div({"class": "num-box"},
                       ui.input_numeric("enr_pval_num", None,
                                        min=0.001, max=1,
                                        value=0.05, step=0.001)),
                col_widths=[8, 4]
            ),

            ui.hr(),

            # ── Report download (Rule 6) ─────────────────────
            ui.tags.h5("Report",
                       style="color:#2C3E50; font-weight:bold;"),
            ui.tags.p(
                "Downloads a .zip folder: HTML report (figures), "
                "DEG CSV, and GO enrichment Excel (all three ontologies).",
                style="font-size:11px; color:#888;"
            ),
            ui.download_button("dl_report", "Download Results Folder"),

            width=300,
        ),

        # ── Main panel ───────────────────────────────────────
        ui.output_ui("validation_msg"),
        ui.output_ui("deg_summary"),
        ui.output_ui("welcome_msg"),
        ui.output_ui("main_tabs"),
    )
)


# ── Server ───────────────────────────────────────────────────
def server(input, output, session):

    # =========================================================
    # Slider ↔ numeric synchronisation
    # Difference from R: observeEvent() → @reactive.effect
    # with @reactive.event(). ignoreInit is handled by
    # checking that the value has actually changed.
    # =========================================================

    @reactive.effect
    @reactive.event(input.deg_pval)
    def _sync_pval_to_num():
        ui.update_numeric("deg_pval_num", value=input.deg_pval())

    @reactive.effect
    @reactive.event(input.deg_pval_num)
    def _sync_pval_to_slider():
        v = input.deg_pval_num()
        if v is not None:
            ui.update_slider("deg_pval", value=v)

    @reactive.effect
    @reactive.event(input.deg_lfc)
    def _sync_lfc_to_num():
        ui.update_numeric("deg_lfc_num", value=input.deg_lfc())

    @reactive.effect
    @reactive.event(input.deg_lfc_num)
    def _sync_lfc_to_slider():
        v = input.deg_lfc_num()
        if v is not None:
            ui.update_slider("deg_lfc", value=v)

    @reactive.effect
    @reactive.event(input.enr_pval)
    def _sync_enr_to_num():
        ui.update_numeric("enr_pval_num", value=input.enr_pval())

    @reactive.effect
    @reactive.event(input.enr_pval_num)
    def _sync_enr_to_slider():
        v = input.enr_pval_num()
        if v is not None:
            ui.update_slider("enr_pval", value=v)


    # =========================================================
    # Rule 3: Reactive data pipeline
    #   raw_data()  →  degs()  →  enr_result()  →  enr_filtered()
    #
    # Difference from R: @reactive.calc replaces reactive().
    # Python has no validate()/req() that propagates to all
    # downstream outputs — each rendering function must check
    # for None explicitly (Rule 4).
    # =========================================================

    # ── Layer 1: Read and validate ───────────────────────────
    @reactive.calc
    def raw_data():
        f = input.file()
        if f is None:
            return None
        try:
            df = pd.read_csv(f[0]["datapath"])
        except Exception as e:
            return None

        required = ["gene", "logFC", "AveExpr", "P.Value", "adj.P.Val"]
        missing  = [c for c in required if c not in df.columns]
        if missing:
            return None      # error shown in validation_msg below

        if df.empty:
            return None
        if not pd.api.types.is_numeric_dtype(df["logFC"]):
            return None
        if not pd.api.types.is_numeric_dtype(df["adj.P.Val"]):
            return None

        return df

    @reactive.calc
    def _missing_cols():
        """Return list of missing columns for validation message."""
        f = input.file()
        if f is None:
            return []
        try:
            df = pd.read_csv(f[0]["datapath"])
        except Exception:
            return ["(could not read file)"]
        required = ["gene", "logFC", "AveExpr", "P.Value", "adj.P.Val"]
        return [c for c in required if c not in df.columns]

    # ── Layer 2: DEG filter ──────────────────────────────────
    @reactive.calc
    def degs():
        df = raw_data()
        if df is None:
            return None
        mask = (
            (df["adj.P.Val"] <= input.deg_pval()) &
            (df["logFC"].abs() >= input.deg_lfc())
        )
        return df[mask].copy()

    # ── Layer 3: GO enrichment (display, one ontology) ───────
    # Difference from R: withProgress() → standard notification.
    # gseapy queries Enrichr — requires internet.
    @reactive.calc
    def enr_result():
        d = degs()
        if d is None or len(d) < 5:
            return None
        gene_set = GO_SETS.get(input.geneset(), GO_SETS["GO_BP"])
        return _run_go(d["gene"].tolist(), gene_set, input.enr_pval())


    # =========================================================
    # Outputs
    # =========================================================

    @output
    @render.ui
    def validation_msg():
        missing = _missing_cols()
        if missing:
            return ui.div(
                f"Missing required columns: {', '.join(missing)}",
                style="color:red; font-weight:bold; padding:8px;"
            )
        return None

    @output
    @render.ui
    def welcome_msg():
        if input.file() is not None:
            return None
        return ui.div(
            {"style": "margin-top:60px; text-align:center; color:#888;"},
            ui.tags.h4("Welcome to the Limma DEG Interactive Explorer",
                       style="color:#2E74B5;"),
            ui.tags.p("To get started, upload your limma output CSV "
                      "using the panel on the left."),
            ui.tags.p(
                {"style": "font-size:13px;"},
                "Required columns: ",
                ui.tags.code("gene"), ", ",
                ui.tags.code("logFC"), ", ",
                ui.tags.code("AveExpr"), ", ",
                ui.tags.code("P.Value"), ", ",
                ui.tags.code("adj.P.Val"),
            ),
            ui.tags.p(
                {"style": "font-size:12px; color:#aaa;"},
                "This is the default output of ",
                ui.tags.code("limma::topTable()"),
                " \u2014 no reformatting needed."
            ),
        )

    @output
    @render.ui
    def main_tabs():
        if input.file() is None:
            return None
        if raw_data() is None:
            return None
        return ui.navset_tab(
            ui.nav_panel("DEG Table",
                         ui.output_data_frame("deg_table")),
            ui.nav_panel("Volcano Plot",
                         ui.output_plot("volcano", height="460px")),
            ui.nav_panel("Top Ranked Genes",
                         ui.output_plot("barplot", height="460px")),
            ui.nav_panel("GO Enrichment",
                         ui.tags.p(
                             "Interactive preview of the selected gene set. "
                             "All three ontologies are in the downloaded Excel. "
                             "Requires internet (Enrichr API). "
                             "May take 10\u201320 seconds.",
                             style="color:#888; font-style:italic; "
                                   "font-size:13px; margin:12px 0;"
                         ),
                         ui.output_plot("enrich_dot", height="420px"),
                         ui.hr(),
                         ui.output_data_frame("enrich_table")),
        )

    @output
    @render.ui
    def deg_summary():
        df = raw_data()
        if df is None:
            return None
        d     = degs()
        n_deg = 0 if d is None else len(d)
        n_up  = 0 if d is None else int((d["logFC"] > 0).sum())
        n_dn  = 0 if d is None else int((d["logFC"] < 0).sum())
        return ui.div(
            {"style": "background:#EBF3FB; padding:10px; border-radius:4px; "
                      "margin-bottom:14px; font-size:14px;"},
            ui.tags.b(f"{n_deg} DEGs"),
            f" from {len(df)} genes  |  ",
            ui.tags.span(f"\u25b2 {n_up} up",
                         style="color:#7B2D8B;"),
            "  ",
            ui.tags.span(f"\u25bc {n_dn} down",
                         style="color:#2E8B57;"),
            ui.tags.span(
                f"adj.P.Val \u2264 {input.deg_pval()}  "
                f"|  |log\u2082FC| \u2265 {input.deg_lfc()}",
                style="color:#888; font-size:12px; margin-left:16px;"
            ),
        )

    # ── Volcano plot ─────────────────────────────────────────
    # Difference from R:
    #   - matplotlib instead of ggplot2
    #   - adjustText instead of ggrepel (approximate non-overlap)
    #   - LaTeX math ($...$) instead of expression() for axis labels
    @output
    @render.plot
    def volcano():
        df = raw_data()
        if df is None:
            return
        df   = df.copy()
        sig  = ((df["adj.P.Val"] <= input.deg_pval()) &
                (df["logFC"].abs() >= input.deg_lfc()))
        df["dir"] = np.where(~sig, "NS",
                    np.where(df["logFC"] > 0, "Up", "Down"))

        # Match R colour scheme: NS=grey, Up=purple, Down=green
        # Legend order matches R: NS -> Up -> Down
        colors      = {"NS": "grey", "Up": "#7B2D8B", "Down": "#2E8B57"}
        plot_order  = ["NS", "Up", "Down"]
        fig, ax = plt.subplots(figsize=(8, 6))

        for grp in plot_order:
            sub = df[df["dir"] == grp]
            sz  = 12 if grp == "NS" else 20
            ax.scatter(sub["logFC"],
                       -np.log10(sub["adj.P.Val"].clip(lower=1e-300)),
                       c=colors[grp], alpha=0.5, s=sz, label=grp, linewidths=0)

        # Threshold lines
        ax.axvline( input.deg_lfc(),  ls="--", color="black", lw=0.7)
        ax.axvline(-input.deg_lfc(),  ls="--", color="black", lw=0.7)
        ax.axhline(-np.log10(input.deg_pval()), ls="--", color="black", lw=0.7)

        # Label top 5 most significant up + top 5 down
        try:
            from adjustText import adjust_text
            top_up   = df[sig & (df["logFC"] > 0)].nsmallest(5, "adj.P.Val")
            top_down = df[sig & (df["logFC"] < 0)].nsmallest(5, "adj.P.Val")
            top_lab  = pd.concat([top_up, top_down])
            texts = []
            for _, row in top_lab.iterrows():
                texts.append(ax.text(
                    row["logFC"],
                    -np.log10(max(row["adj.P.Val"], 1e-300)),
                    row["gene"], fontsize=7
                ))
            adjust_text(texts, ax=ax,
                        arrowprops=dict(arrowstyle="-", color="grey", lw=0.5))
        except ImportError:
            pass

        ax.set_xlabel(r"$\log_2$ Fold Change", fontsize=11)
        ax.set_ylabel(r"$-\log_{10}$(adj. P-value)", fontsize=11)
        ax.set_title("Volcano Plot", fontsize=13, fontweight="bold", pad=12)
        ax.legend(title=None, frameon=False, loc="upper left", markerscale=1.2, fontsize=9)
        plt.tight_layout()
        return fig

    # ── Top ranked genes bar plot ────────────────────────────
    @output
    @render.plot
    def barplot():
        d = degs()
        if d is None or d.empty:
            return
        top = d.copy()
        top["rank"]      = np.sign(top["logFC"]) * -np.log10(top["adj.P.Val"])
        top["direction"] = np.where(top["logFC"] > 0, "Up", "Down")
        top = top.reindex(top["rank"].abs().sort_values(ascending=False).index
                          ).head(30)

        colors = ["#7B2D8B" if v > 0 else "#2E8B57"
                  for v in top["rank"]]
        fig, ax = plt.subplots(figsize=(9, 10))
        ax.barh(range(len(top)), top["rank"].values, color=colors)
        ax.set_yticks(range(len(top)))
        ax.set_yticklabels(top["gene"].values, fontsize=8)
        ax.set_xlabel(
            r"Rank score: $\mathrm{sign}(\log_2 FC) \times -\log_{10}(\mathrm{adj.\ P.Val})$",
            fontsize=9, labelpad=10
        )
        ax.set_title(
            "Top 30 DEGs by rank",
            fontsize=11, pad=10
        )
        ax.axvline(0, color="black", lw=0.5)

        from matplotlib.patches import Patch
        ax.legend(
            handles=[
                Patch(color="#7B2D8B", label="Up"),
                Patch(color="#2E8B57", label="Down")
            ],
            frameon=False,
            loc="upper left",
            bbox_to_anchor=(1.01, 1),
            borderaxespad=0
        )

        plt.tight_layout()
        fig.subplots_adjust(right=0.88)
        return fig

    # ── DEG table ────────────────────────────────────────────
    @output
    @render.data_frame
    def deg_table():
        d = degs()
        if d is None or d.empty:
            return pd.DataFrame()
        cols = [c for c in ["gene","logFC","AveExpr","P.Value","adj.P.Val","B"]
                if c in d.columns]
        return d[cols].sort_values("adj.P.Val")

    # ── GO enrichment dot plot ───────────────────────────────
    @output
    @render.plot
    def enrich_dot():
        ef = enr_result()
        if ef is None or ef.empty:
            fig, ax = plt.subplots(figsize=(7, 3))
            ax.text(0.5, 0.5,
                    "No enrichment terms at current cutoff.\n"
                    "(Requires internet connection)",
                    ha="center", va="center", fontsize=11, color="#888",
                    transform=ax.transAxes)
            ax.axis("off")
            return fig

        top = ef.nsmallest(15, "p.adjust")

        # Scale dot size by Count to match R clusterProfiler dot plot
        count_min = top["Count"].min()
        count_max = top["Count"].max()
        count_range = max(count_max - count_min, 1)
        dot_sizes = 50 + 300 * (top["Count"] - count_min) / count_range

        fig, ax = plt.subplots(figsize=(10, 7))
        sc = ax.scatter(
            top["Count"],
            range(len(top)),
            c=top["p.adjust"],
            cmap="RdYlBu_r",
            s=dot_sizes,
            edgecolors="none"
        )
        ax.set_yticks(range(len(top)))
        ax.set_yticklabels(top["Description"].values, fontsize=9)
        ax.set_xlabel("Gene count", fontsize=11)
        ax.set_title(
            f"Top 15 terms \u2014 {input.geneset()} "
            f"(adj. P-value \u2264 {input.enr_pval()})",
            fontsize=12, fontweight="bold", pad=12
        )

        # Colorbar matching R "adj. P-value" label
        cbar = plt.colorbar(sc, ax=ax, shrink=0.6, pad=0.12)
        cbar.set_label("adj. P-value", fontsize=9)

        # Count size legend matching R style
        from matplotlib.lines import Line2D
        count_vals = sorted(set(
            [count_min,
             count_min + (count_max - count_min) / 3,
             count_min + 2 * (count_max - count_min) / 3,
             count_max]
        ))
        count_vals = [int(round(v)) for v in count_vals]
        legend_elements = [
            Line2D([0], [0], marker="o", color="w",
                   markerfacecolor="grey",
                   markersize=(50 + 300 * (v - count_min) / count_range) ** 0.5,
                   label=str(v))
            for v in count_vals
        ]
        ax.legend(
            handles=legend_elements,
            title="Count", frameon=False,
            loc="upper left",
            bbox_to_anchor=(1.15, 1),
            borderaxespad=0,
            fontsize=8, title_fontsize=9
        )

        plt.tight_layout()
        fig.subplots_adjust(right=0.75)
        return fig

    # ── GO enrichment table ──────────────────────────────────
    @output
    @render.data_frame
    def enrich_table():
        ef = enr_result()
        if ef is None or ef.empty:
            return pd.DataFrame({"Message": [
                "No enrichment terms — check internet connection "
                "or adjust the adj. P-value cutoff."
            ]})
        cols = [c for c in ["Description","pvalue","p.adjust","Count","geneID"]
                if c in ef.columns]
        return ef[cols].sort_values("p.adjust")


    # =========================================================
    # Rule 6: Download ZIP folder
    #
    # Contents (mirrors R implementation exactly):
    #   DEG_report_TIMESTAMP.html    — figures only (Jinja2)
    #   DEG_results_TIMESTAMP.csv    — filtered DEG table
    #   GO_enrichment_TIMESTAMP.xlsx — GO_BP, GO_MF, GO_CC sheets
    #
    # Difference from R:
    #   R uses rmarkdown::render() + openxlsx + zip().
    #   Python uses Jinja2 + base64 figures + openpyxl + zipfile.
    # =========================================================

    def _make_volcano_fig():
        """Re-render volcano plot for the report."""
        df = raw_data()
        if df is None:
            return plt.figure()
        df   = df.copy()
        sig  = ((df["adj.P.Val"] <= input.deg_pval()) &
                (df["logFC"].abs() >= input.deg_lfc()))
        df["dir"] = np.where(~sig, "NS",
                    np.where(df["logFC"] > 0, "Up", "Down"))
        colors = {"NS": "grey", "Up": "#7B2D8B", "Down": "#2E8B57"}
        fig, ax = plt.subplots(figsize=(8, 5))
        for grp, color in colors.items():
            sub = df[df["dir"] == grp]
            ax.scatter(sub["logFC"],
                       -np.log10(sub["adj.P.Val"].clip(lower=1e-300)),
                       c=color, alpha=0.5, s=12, label=grp, linewidths=0)
        ax.axvline( input.deg_lfc(), ls="--", color="black", lw=0.7)
        ax.axvline(-input.deg_lfc(), ls="--", color="black", lw=0.7)
        ax.axhline(-np.log10(input.deg_pval()), ls="--", color="black", lw=0.7)
        try:
            from adjustText import adjust_text
            top_up   = df[sig & (df["logFC"] > 0)].nsmallest(5, "adj.P.Val")
            top_down = df[sig & (df["logFC"] < 0)].nsmallest(5, "adj.P.Val")
            top_lab  = pd.concat([top_up, top_down])
            texts = [ax.text(r["logFC"],
                             -np.log10(max(r["adj.P.Val"], 1e-300)),
                             r["gene"], fontsize=7)
                     for _, r in top_lab.iterrows()]
            adjust_text(texts, ax=ax,
                        arrowprops=dict(arrowstyle="-", color="grey", lw=0.5))
        except ImportError:
            pass
        ax.set_xlabel(r"$\log_2$ Fold Change")
        ax.set_ylabel(r"$-\log_{10}$(adj. P-value)")
        ax.set_title("Volcano Plot")
        ax.legend(frameon=False)
        plt.tight_layout()
        return fig

    def _make_barplot_fig():
        """Re-render bar plot for the report."""
        d = degs()
        if d is None or d.empty:
            return plt.figure()
        top = d.copy()
        top["rank"] = np.sign(top["logFC"]) * -np.log10(top["adj.P.Val"])
        top = top.reindex(top["rank"].abs().sort_values(ascending=False).index
                          ).head(30)
        colors = ["#7B2D8B" if v > 0 else "#2E8B57" for v in top["rank"]]
        fig, ax = plt.subplots(figsize=(7, 8))
        ax.barh(range(len(top)), top["rank"].values, color=colors)
        ax.set_yticks(range(len(top)))
        ax.set_yticklabels(top["gene"].values, fontsize=8)
        ax.set_xlabel(
            r"Rank: $\mathrm{sign}(\log_2 FC) \times -\log_{10}(\mathrm{adj.P.Val})$",
            fontsize=9
        )
        ax.set_title("Top 30 DEGs by rank", fontsize=10)
        ax.axvline(0, color="black", lw=0.5)
        plt.tight_layout()
        return fig

    def _make_go_fig(df, title):
        """Render a GO dot plot for the report — matches R clusterProfiler style."""
        if df is None or df.empty:
            return None
        top = df.nsmallest(15, "p.adjust")

        count_min  = top["Count"].min()
        count_max  = top["Count"].max()
        count_range = max(count_max - count_min, 1)
        dot_sizes  = 50 + 300 * (top["Count"] - count_min) / count_range

        fig, ax = plt.subplots(figsize=(10, 7))
        sc = ax.scatter(top["Count"], range(len(top)),
                        c=top["p.adjust"], cmap="RdYlBu_r",
                        s=dot_sizes, edgecolors="none")
        ax.set_yticks(range(len(top)))
        ax.set_yticklabels(top["Description"].values, fontsize=9)
        ax.set_xlabel("Gene count", fontsize=11)
        ax.set_title(title, fontsize=12, fontweight="bold", pad=12)

        cbar = plt.colorbar(sc, ax=ax, shrink=0.6, pad=0.12)
        cbar.set_label("adj. P-value", fontsize=9)

        from matplotlib.lines import Line2D
        count_min  = top["Count"].min()
        count_vals = [int(round(v)) for v in [
            count_min,
            count_min + (count_max - count_min) / 3,
            count_min + 2 * (count_max - count_min) / 3,
            count_max
        ]]
        legend_elements = [
            Line2D([0], [0], marker="o", color="w",
                   markerfacecolor="grey",
                   markersize=(50 + 300 * (v - count_min) / count_range) ** 0.5,
                   label=str(v))
            for v in count_vals
        ]
        ax.legend(handles=legend_elements, title="Count", frameon=False,
                  loc="upper left", bbox_to_anchor=(1.15, 1),
                  borderaxespad=0, fontsize=8, title_fontsize=9)

        plt.tight_layout()
        fig.subplots_adjust(right=0.75)
        return fig

    def _write_go_sheet(ws, df, sheet_title):
        """Write a GO enrichment DataFrame to an openpyxl worksheet."""
        go_cols = [c for c in ["Description","pvalue","p.adjust","Count","geneID"]
                   if c in df.columns]
        df = df[go_cols].sort_values("p.adjust")

        # Header style
        hdr_fill = PatternFill("solid", fgColor="2E74B5")
        hdr_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
        hdr_aln  = Alignment(horizontal="left")

        for ci, col in enumerate(go_cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_aln

        # Data rows
        data_font = Font(name="Arial", size=10)
        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            for ci, col in enumerate(go_cols, start=1):
                cell = ws.cell(row=ri, column=ci, value=row[col])
                cell.font = data_font

        # Auto-width
        for ci, col in enumerate(go_cols, start=1):
            max_len = max(len(str(col)),
                          df[col].astype(str).str.len().max()
                          if not df.empty else 0)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)

    @render.download(
        filename=lambda: (
            f"DEG_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
        )
    )
    async def dl_report():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        d  = degs()

        # ── 1. DEG CSV ──────────────────────────────────────
        deg_cols = [c for c in ["gene","logFC","AveExpr","P.Value","adj.P.Val","B"]
                    if c in d.columns]
        csv_bytes = d[deg_cols].sort_values("adj.P.Val").to_csv(index=False
                                                                 ).encode()

        # ── 2. All three GO enrichments ─────────────────────
        gene_symbols = d["gene"].tolist()
        enr_bp = _run_go(gene_symbols, GO_SETS["GO_BP"], input.enr_pval())
        enr_mf = _run_go(gene_symbols, GO_SETS["GO_MF"], input.enr_pval())
        enr_cc = _run_go(gene_symbols, GO_SETS["GO_CC"], input.enr_pval())

        # ── 3. GO enrichment Excel ──────────────────────────
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        sheets_added = 0
        for sheet_name, df_go in [("GO_BP", enr_bp),
                                   ("GO_MF", enr_mf),
                                   ("GO_CC", enr_cc)]:
            if df_go is not None and not df_go.empty:
                ws = wb.create_sheet(sheet_name)
                _write_go_sheet(ws, df_go, sheet_name)
                sheets_added += 1

        if sheets_added == 0:
            ws = wb.create_sheet("No_results")
            ws.cell(1, 1,
                    f"No enriched terms at adj. P-value ≤ {input.enr_pval()}")

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_bytes = xlsx_buf.getvalue()

        # ── 4. HTML report (figures only, Jinja2) ───────────
        # Difference from R: R Markdown renders ggplot2 objects.
        # Here, matplotlib figures are converted to base64 PNG.
        vol_b64  = _fig_to_b64(_make_volcano_fig())
        bar_b64  = _fig_to_b64(_make_barplot_fig())
        go_bp_fig = _make_go_fig(
            enr_bp,
            f"GO Biological Process (adj. P-value ≤ {input.enr_pval()})"
        )
        go_mf_fig = _make_go_fig(
            enr_mf,
            f"GO Molecular Function (adj. P-value ≤ {input.enr_pval()})"
        )
        go_cc_fig = _make_go_fig(
            enr_cc,
            f"GO Cellular Component (adj. P-value ≤ {input.enr_pval()})"
        )

        tmpl_path = os.path.join(
            os.path.dirname(__file__), "templates", "report_template.html"
        )
        with open(tmpl_path) as f:
            tmpl = Template(f.read())

        html_str = tmpl.render(
            date         = datetime.now().strftime("%d %B %Y, %H:%M"),
            deg_pval     = input.deg_pval(),
            deg_lfc      = input.deg_lfc(),
            enr_pval     = input.enr_pval(),
            n_total      = len(raw_data()),
            n_degs       = len(d),
            n_up         = int((d["logFC"] > 0).sum()),
            n_dn         = int((d["logFC"] < 0).sum()),
            n_bp         = len(enr_bp) if enr_bp is not None else 0,
            n_mf         = len(enr_mf) if enr_mf is not None else 0,
            n_cc         = len(enr_cc) if enr_cc is not None else 0,
            volcano_b64  = vol_b64,
            barplot_b64  = bar_b64,
            go_bp_b64    = _fig_to_b64(go_bp_fig) if go_bp_fig else None,
            go_mf_b64    = _fig_to_b64(go_mf_fig) if go_mf_fig else None,
            go_cc_b64    = _fig_to_b64(go_cc_fig) if go_cc_fig else None,
        )
        html_bytes = html_str.encode()

        # ── 5. Bundle into ZIP ──────────────────────────────
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"DEG_report_{timestamp}.html",    html_bytes)
            zf.writestr(f"DEG_results_{timestamp}.csv",    csv_bytes)
            zf.writestr(f"GO_enrichment_{timestamp}.xlsx", xlsx_bytes)

        yield zip_buf.getvalue()


# ── Launch ───────────────────────────────────────────────────
app = App(app_ui, server)
